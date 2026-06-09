"""
05_imu_convert.py

Converts Xsens MVN Awinda Excel exports to CSV for each trial.
Extracts only the Joint Angles ZXY sheet, which contains the joint
angles in degrees already computed by the Xsens firmware.

Joint angles extracted (8 signals):
  - Right/Left Hip Flexion/Extension
  - Right/Left Hip Abduction/Adduction
  - Right/Left Knee Flexion/Extension
  - Right/Left Ankle Dorsiflexion/Plantarflexion

The analysis window (frames 1200-3000, seconds 20-50) is NOT applied
here. It is applied in 06_imu_pipeline.py to keep conversion and
analysis separate.

Output:
  imu_csv/<participant>/trial-XXX_joint_angles.csv

Usage:
    python 05_imu_convert.py
"""

import os
import glob
import openpyxl
import pandas as pd

# ══════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════
PARTICIPANTS = ["p2", "p1", "p3", "p4"]

# Columns to extract from 'Joint Angles ZXY'
COLS = [
    "Frame",
    "Right Hip Abduction/Adduction",
    "Right Hip Flexion/Extension",
    "Left Hip Abduction/Adduction",
    "Left Hip Flexion/Extension",
    "Right Knee Flexion/Extension",
    "Left Knee Flexion/Extension",
    "Right Ankle Dorsiflexion/Plantarflexion",
    "Left Ankle Dorsiflexion/Plantarflexion",
]

# Short names for CSV output
COL_RENAME = {
    "Right Hip Abduction/Adduction":          "R_hip_abd",
    "Right Hip Flexion/Extension":            "R_hip_flex",
    "Left Hip Abduction/Adduction":           "L_hip_abd",
    "Left Hip Flexion/Extension":             "L_hip_flex",
    "Right Knee Flexion/Extension":           "R_knee_flex",
    "Left Knee Flexion/Extension":            "L_knee_flex",
    "Right Ankle Dorsiflexion/Plantarflexion":"R_ankle_dors",
    "Left Ankle Dorsiflexion/Plantarflexion": "L_ankle_dors",
}

# ══════════════════════════════════════════════════════════════════
# PATHS
# ══════════════════════════════════════════════════════════════════
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "..")          # TFG/Data/
OUT_DIR  = os.path.join(BASE_DIR, "imu_csv")
os.makedirs(OUT_DIR, exist_ok=True)

# ══════════════════════════════════════════════════════════════════
# CONVERSION
# ══════════════════════════════════════════════════════════════════
for participant in PARTICIPANTS:
    p_cap    = participant.capitalize()
    src_dir  = os.path.join(DATA_DIR, f"{p_cap}_data")
    out_pdir = os.path.join(OUT_DIR, participant)

    xlsx_files = sorted(glob.glob(os.path.join(src_dir, "trial-*.xlsx")))

    if not xlsx_files:
        print(f"[SKIP] {participant}: no s'han trobat fitxers trial-XXX.xlsx")
        continue

    os.makedirs(out_pdir, exist_ok=True)
    print(f"\n{'='*50}")
    print(f"  {participant.upper()}  ({len(xlsx_files)} trials)")
    print(f"{'='*50}")

    for xlsx_path in xlsx_files:
        trial_name = os.path.splitext(os.path.basename(xlsx_path))[0]  # 'trial-001'

        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

            if "Joint Angles ZXY" not in wb.sheetnames:
                print(f"  [WARN] {trial_name}: pestanya 'Joint Angles ZXY' no trobada")
                continue

            ws   = wb["Joint Angles ZXY"]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            header = list(rows[0])
            data   = rows[1:]
            df_raw = pd.DataFrame(data, columns=header)

            # Select and rename columns
            df = df_raw[COLS].rename(columns=COL_RENAME)
            df = df.reset_index(drop=True)

            # Save
            out_path = os.path.join(out_pdir, f"{trial_name}_joint_angles.csv")
            df.to_csv(out_path, index=False)

            print(f"  ✓ {trial_name}  →  {df.shape[0]} frames × {df.shape[1]} cols")

        except Exception as e:
            print(f"  [ERROR] {trial_name}: {e}")

print(f"\n{'='*50}")
print(f"✓ CSVs guardats a: {OUT_DIR}")
print(f"\nAra executa: python 06_imu_pipeline.py")
